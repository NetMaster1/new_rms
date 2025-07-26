from multiprocessing.connection import Client
from shutil import move
from ssl import create_default_context
from app_personnel.models import BonusAccount, Salary
from django.db.models import query
from app_product.views import change_cash_receipt_unposted
from app_reference.models import DocumentType, ProductCategory, Shop, Supplier, Product, Expense, Month, Year
from app_cash.models import Cash, Credit, Card
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages, auth
from django.contrib.auth.models import User, Group
from django.contrib.auth import logout, login
from app_product.models import (
    Product,
    RemainderHistory,
    RemainderCurrent,
    Document,
)
from app_kpi.models import KPIMonthlyPlan


from .models import (
    ReportTemp,
    ReportTempId,
    DailySaleRep,
    MonthlyBonus,
    SaleReport,
    PayCardReport,
    ClientReport,
    ClientHistoryReport,
    AcquiringReport,
    DeliveryReport,
    ExpensesReport,
    SalaryReport,
    EffectivenessReport,
)
from app_clients.models import Customer
from app_personnel.models import BulkSimMotivation
from .models import ProductHistory
from django.contrib import messages
import pandas as pd
import numpy as np
from datetime import datetime, date, timedelta
import xlwt
import openpyxl
from openpyxl import Workbook, load_workbook
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator

# import xlwings as xw
# import pro
from django.http import HttpResponse, JsonResponse
import datetime
import pytz
import os
from django.db.models import Q

# Create your views here.

def daily_report(request):
    return render(request, "reports/daily_report.html")

def save_in_excel_daily_rep(request):
    categories = ProductCategory.objects.all().order_by('id')
    #shops = Shop.objects.all().order_by("name").exclude(name="ООС")
    shops = Shop.objects.filter(active=True, retail=True).order_by("name")
    # length=shops.count()
    doc_type = DocumentType.objects.get(name="Продажа ТМЦ")
    pay_type = DocumentType.objects.get(name="Платежи Теко")
    if request.method == "POST":
        date = request.POST.get("date", False)
        if date:
            # converting HTML date format (2021-07-08T01:05) to django format (2021-07-10 01:05:00)
            date = datetime.datetime.strptime(date, "%Y-%m-%d")
        else:
            date = datetime.date.today()
        dateTime = date.strftime("%Y-%m-%d")

        report_id = ReportTempId.objects.create()
        for shop in shops:
            shop_row = []  # list for storing retail values for each category/shop for further placing in model
            for category in categories:
                sum = 0
                if RemainderHistory.objects.filter(shop=shop, category=category, created__date=date, rho_type=doc_type).exists():
                    rhos = RemainderHistory.objects.filter(shop=shop, category=category, created__date=date, rho_type=doc_type)
                    for rho in rhos:
                        sum += rho.sub_total
                else:
                    sum=0
                shop_row.append(sum)
            if Cash.objects.filter(shop=shop, cho_type=pay_type, created__date=date).exists():
                chos=Cash.objects.filter(shop=shop, cho_type=pay_type, created__date=date)
                for cho in chos:
                    sum+=cho.cash_in
            else:
                sum=0
            shop_row.append(sum)
            # ===================================================================================
            expenses_type = DocumentType.objects.get(name="РКО (хоз.расходы)")
            if Cash.objects.filter(cho_type=expenses_type, shop=shop, created__date=date).exists():
                chos = Cash.objects.filter(cho_type=expenses_type, shop=shop, created__date=date)
                expenses_sum = 0
                for cho in chos:
                    expenses_sum += cho.cash_out
            else:
                expenses_sum = 0
            salary_type = DocumentType.objects.get(name="РКО (зарплата)")
            if Cash.objects.filter(cho_type=salary_type, shop=shop, created__date=date).exists():
                chos = Cash.objects.filter(cho_type=salary_type, shop=shop, created__date=date)
                salary_sum = 0
                for cho in chos:
                    salary_sum += cho.cash_out
            else:
                salary_sum = 0
            cash_move_type = DocumentType.objects.get(name="Перемещение денег")
            if Cash.objects.filter(cho_type=cash_move_type, shop=shop, sender=True, created__date=date).exists():
                chos = Cash.objects.filter(cho_type=cash_move_type, shop=shop, sender=True, created__date=date)
                cash_move_sum = 0
                for cho in chos:
                    cash_move_sum += cho.cash_out
            else:
                cash_move_sum = 0
            return_type = DocumentType.objects.get(name="Возврат ТМЦ")
            if Cash.objects.filter(cho_type=return_type, shop=shop, created__date=date).exists():
                chos = Cash.objects.filter(cho_type=return_type, shop=shop, created__date=date)
                return_sum = 0
                for cho in chos:
                    return_sum += cho.cash_out
            else:
                return_sum = 0
            if Card.objects.filter(shop=shop, created__date=date).exists():
                cards = Card.objects.filter(shop=shop, created__date=date)
                card_sum = 0
                for card in cards:
                    card_sum += card.sum
            else:
                card_sum = 0
            if Credit.objects.filter(shop=shop, created__date=date).exists():
                credits = Credit.objects.filter(shop=shop, created__date=date)
                credit_sum = 0
                for credit in credits:
                    credit_sum += credit.sum
            else:
                credit_sum = 0
            #=========================Calculating Cashback=============================
            sale_type = DocumentType.objects.get(name="Продажа ТМЦ")
            if Document.objects.filter(shop_sender=shop, created__date=date, posted=True, title=sale_type).exists():
                docs=Document.objects.filter(shop_sender=shop, created__date=date, posted=True, title=sale_type)
                cashback=0
                for doc in docs:
                    cashback+=doc.cashback_off
            else:
                cashback=0
            # =======================Calculating opening balance for each shop======================
            if Cash.objects.filter(created__lt=date, shop=shop).exists():
                prev_day_cho = Cash.objects.filter(created__lt=date, shop=shop).latest("created")
                prev_day_cash_remainder = prev_day_cho.current_remainder
            else:
                prev_day_cash_remainder = 0
            # ================================================
            daily_rep = DailySaleRep.objects.create(
                report_id=report_id,
                created=dateTime,
                shop=shop.name,
                opening_balance=prev_day_cash_remainder,
                smartphones=shop_row[0],
                accessories=shop_row[1],
                sim_cards=shop_row[2],
                phones=shop_row[3],
                iphone=shop_row[4],
                insuranсе=shop_row[5],
                wink=shop_row[6],
                services=shop_row[7],
                pay_cards=shop_row[8],
                gadgets=shop_row[9],
                modems=shop_row[10],
                RT_equipment=shop_row[11],
                teko_payments=shop_row[12],
                protective_films=shop_row[13],
                credit=credit_sum,
                card=card_sum,
                cashback=cashback,
                salary=salary_sum,
                expenses=expenses_sum,
                return_sum=return_sum,
                cash_move=cash_move_sum,
            )
            daily_rep.net_sales = (
                #listing categorie must be as they go in the existin DB
                daily_rep.smartphones
                + daily_rep.accessories
                + daily_rep.sim_cards
                + daily_rep.phones
                + daily_rep.iphone
                + daily_rep.insuranсе
                + daily_rep.wink
                + daily_rep.services
                + daily_rep.pay_cards
                + daily_rep.gadgets
                + daily_rep.modems
                + daily_rep.RT_equipment
                + daily_rep.protective_films
                + daily_rep.teko_payments
                
                
            )
            daily_rep.final_balance = (
                daily_rep.opening_balance
                + daily_rep.net_sales
                - daily_rep.credit
                - daily_rep.salary
                - daily_rep.card
                - daily_rep.cashback
                - daily_rep.expenses
                - daily_rep.return_sum
                - daily_rep.cash_move
            )
            daily_rep.save()
         
#==========================Convert to Excel module=========================================
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = (
            "attachment; filename=DailRep_" + str(date) + ".xls"
        )
        # str(datetime.date.today())+'.xls'

        wb = xlwt.Workbook(encoding="utf-8")
        ws = wb.add_sheet(dateTime)

        # sheet header in the first row
        row_num = 0
        font_style = xlwt.XFStyle()

        columns = []
        for shop in shops:
            columns.append(shop.name)
        for col_num in range(len(columns)):
            ws.write(row_num, col_num + 1, columns[col_num], font_style)

        # sheet body, remaining rows
        font_style = xlwt.XFStyle()
        daily_rep = DailySaleRep.objects.filter(report_id=report_id)

        col_num = 1
        for shop in shops:
            query_list = daily_rep.get(shop=shop.name)
            # query = query_list.values_list("opening_balance", "smartphones", "accessories", 'sim_cards', 'phones', 'iphone', 
            #'insuranсе', 'wink', 'services', 'modems', 'RT_equipment', 'protective_films', pay_cards', 'credit', 'card', 'salary', 'expenses', 'return_sum', 'cash_move', 'final_balance')
            row_num = 1
            ws.write(row_num, col_num, query_list.opening_balance, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.smartphones, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.accessories, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.sim_cards, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.phones, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.iphone, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.insuranсе, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.wink, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.services, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.pay_cards, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.gadgets, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.modems, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.RT_equipment, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.protective_films, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.teko_payments, font_style)
            

            row_num += 1
            ws.write(row_num, col_num, query_list.credit, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.card, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.cashback, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.salary, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.expenses, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.return_sum, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.cash_move, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.final_balance, font_style)
            col_num += 1

        criteria_list = [
            "Остаток на начало смены",
            "Смартфоны",
            "Аксы",
            "Сим_карты",
            "Трубки",
            "Iphone",
            "Страховка",
            "wink",
            "услуги",
            "КЭО",
            "Гаджеты",
            "Модемы",
            "Оборудование РТ",
            "Пленки",
            "Платежи",
            "Продажа в кредит",
            "Экваиринг",
            "Кэшбэк",
            "Зар. плата",
            "Расходы",
            "Возврат",
            "Перемещение ден. средств",
            "Остаток на конец смены",
        ]
        row_num = 1
        for i in criteria_list:
            ws.write(row_num, 0, i, font_style)
            row_num = row_num + 1

        wb.save(response)
        return response

#=================CashBack Report================================

def clients_per_user(request):
    if request.user.is_authenticated:
        users=User.objects.all().order_by('last_name')
        if request.method == 'POST':
            start_date=request.POST ['start_date']
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            end_date = request.POST ["end_date"]
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            end_date = end_date + timedelta(days=1)
            user = request.POST["user"]
            user=User.objects.get(id=user)
            if Customer.objects.filter(user=user, created__gte=start_date, created__lt=end_date).exists():
                customers=Customer.objects.filter (user=user, created__gte=start_date, created__lt=end_date)
            else:
                messages.error(request, "Новые клиенты отсутствуют")
                return redirect("clients_per_user")
        else:
            context = {
                'users': users
            }
            return render (request, "reports/cashback_rep.html", context)

    else:
        auth.logout(request)
        return redirect("login")

def cashback_rep (request):
    if request.user.is_authenticated:
        users=User.objects.all().order_by('last_name')
        if request.method == 'POST':
            start_date=request.POST ['start_date']
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            end_date = request.POST ["end_date"]
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            end_date = end_date + timedelta(days=1)
            print(start_date)
            print(end_date)
            if Customer.objects.filter(created__gte=start_date, created__lt=end_date).exists():#look for new clients who were created in this time period
                customers=Customer.objects.filter (created__gte=start_date, created__lt=end_date)
            else:
                messages.error(request, "Новые клиенты отсутствуют")
                return redirect("cashback_rep")
            doc_type=DocumentType.objects.get(name='Продажа ТМЦ')
            if Document.objects.filter(created__gte=start_date, created__lt=end_date, title=doc_type).exists():
                documents=Document.objects.filter(created__gte=start_date, created__lt=end_date, title=doc_type)
            else:
                messages.error(request, "Продаж в данный период не было")
                return redirect("cashback_rep")
            dict = {}
            for user in users:
                counter=0
                for customer in customers:
                    if Document.objects.filter(client=customer).exists():#we add 1 to the counter even if the customer exists in more than one document
                        if customer.user == user:
                            counter +=1
                if counter > 0:
                    dict[user]=counter

            context = {
                'dict': dict
            }

            return render (request, "reports/cashback_rep.html", context)
        else:
            context = {
                'users': users
            }
            return render (request, "reports/cashback_rep.html", context)
    else:
        auth.logout(request)
        return redirect("login")

def cashback_history (request):
    group_sales=Group.objects.get(name='sales')
    users = User.objects.filter(is_active=True, groups=group_sales ).order_by('username')
    if request.method=="POST":
        doc_type=DocumentType.objects.get(name='Продажа ТМЦ')
        user = request.POST["user"]
        user=User.objects.get(id=user)
        start_date = request.POST["start_date"]
        start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
        end_date = request.POST["end_date"]
        end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
        end_date = end_date + timedelta(days=1)
        if Document.objects.filter(title=doc_type, user=user, created__gt=start_date, created__lt=end_date).exists():
            documents_general=Document.objects.filter(title=doc_type, user=user, created__gt=start_date, created__lt=end_date)
            documents_total=Document.objects.filter(title=doc_type, created__gt=start_date, created__lt=end_date)
        else:
            messages.error(request, "У данного пользователя не было продаж в данный период.")
            return redirect("cashback_history")
        clients=Customer.objects.all().exclude(phone='79200711112')
        report_id=ReportTempId.objects.create()

        for client in clients:
            if documents_general.filter(client=client).exists():
                documents=documents_general.filter(client=client)
                number=documents.count()

                client_rep=ClientReport.objects.create(
                    report_id=report_id,
                    phone=client.phone,
                    user=user,
                    created=client.created,
                    count=number
                )
                cashback_off=0
                cashback_awarded=0
                for document in documents:
                    cashback_off+=document.cashback_off
                client_rep.cashback_off=cashback_off
                # client_rep.save()
                for document in documents:
                    rhos=RemainderHistory.objects.filter(document=document)
                    for rho in rhos:
                        if rho.cash_back_awarded is not None:
                            cashback_awarded+=rho.cash_back_awarded
                client_rep.cashback_awarded=cashback_awarded
                # client_rep.save()
                client_rep.cashback_remaining=client.accum_cashback
                client_rep.save()
               
        queryset_list=ClientReport.objects.filter(report_id=report_id).order_by('-count')
        # ============paginator module=================
        # paginator = Paginator(clients, 50)
        # page = request.GET.get('page')
        # queryset_list = paginator.get_page(page)
        # =============end of paginator module===============
        context = {
            'documents_general': documents_general,
            'documents_total': documents_total,
            'queryset_list': queryset_list,
            'start_date': start_date,
            'end_date': end_date,
            'users': users
        }
        return render (request, 'reports/cashback_history.html', context)
    else:
        context = {
            'users': users
        }
        return render (request, 'reports/cashback_history.html', context)

def clients_history_report(request):
    clients=Customer.objects.all()
    doc_type = DocumentType.objects.get(name="Продажа ТМЦ")
    report_id=ReportTempId.objects.create()
    #if request.method=="POST":
    for item in clients:
        counter=0
        cashback_off=0
        cashback_awarded=0
        sum=0
        docs=Document.objects.filter(client=item)     
        for doc in docs:
            rhos=RemainderHistory.objects.filter(document=doc)
            for rho in rhos:
                if rho.cash_back_awarded is not None:
                    cashback_awarded+=rho.cash_back_awarded
            if doc.sum is not None:
                sum+=doc.sum
            else:
                sum+=0
            cashback_off+=doc.cashback_off
            counter+=1
        client=ClientHistoryReport.objects.create(
            report_id=report_id,
            phone=item.phone,
            number_of_docs=counter,
            sum=sum,
            cashback_off=cashback_off,
            cashback_awarded=cashback_awarded,
            user=item.user
        )
    queryset_list=ClientHistoryReport.objects.filter(report_id=report_id).order_by('-number_of_docs')
    qs=queryset_list.values('phone', 'user', 'number_of_docs', 'sum', 'cashback_awarded', 'cashback_off')
    data=pd.DataFrame.from_records(qs)
    data.to_excel('clientsHistory.xlsx')
    os.system('start excel.exe clientsHistory.xlsx')
    return render(request, "reports/clients_history_report.html")

    #============paginator module=================
    # paginator = Paginator(queryset_list, 200)
    # page = request.GET.get('page')
    # paged_queryset_list = paginator.get_page(page)
    # #=============end of paginator module===============
    # context = {
    #     'queryset_list': paged_queryset_list
    # }
    # return render (request, 'reports/clients_history_report.html', context)
# ===============================================================

def delete_clients_history_report (request):
    items=ClientHistoryReport.objects.all()
    for item in items:
        item.delete()
    return redirect ('log')

def close_report(request):
    if request.user.is_authenticated:
        users = Group.objects.get(name="sales").user_set.all()
        # if ReportTemp.objects.filter(existance_check=True).exists():
        #     reports_temp = ReportTemp.objects.all()
        #     for obj in reports_temp:
        #         obj.delete()
        # if ReportTempId.objects.filter(existance_check=True).exists():
        #     report_ids_temp = ReportTempId.objects.all()
        #     for obj in report_ids_temp:
        #         obj.delete()
        if request.user in users:
            return redirect("sale_interface")
        else:
            return redirect("log")
    else:
        auth.logout(request)
        return redirect("login")

def close_remainder_report(request):
    users = Group.objects.get(name="sales").user_set.all()
    if request.user in users:
        return redirect("sale_interface")
    else:
        return redirect("log")

#==================================================================
def sale_report_per_shop(request):
    shops = Shop.objects.all()
    if request.method == "POST":
        doc_type = DocumentType.objects.get(name="Продажа ТМЦ")
        teko_cash_in=DocumentType.objects.get(name="Платежи Теко")
        salary_cash_out=DocumentType.objects.get(name="РКО (зарплата)")
        expenses_cash_out=DocumentType.objects.get(name="РКО (хоз.расходы)")
        money_transfer_cash_out=DocumentType.objects.get(name="Перемещение денег")
        money_return_cash_out=DocumentType.objects.get(name="Возврат ТМЦ")
        shop = request.POST["shop"]
        shop = Shop.objects.get(id=shop)
        start_date = request.POST["start_date"]
        start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
        end_date = request.POST["end_date"]
        end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
        end_date = end_date + timedelta(days=1)
        if RemainderHistory.objects.filter(rho_type=doc_type, shop=shop, created__gt=start_date, created__lt=end_date).exists():
            queryset =RemainderHistory.objects.filter(rho_type=doc_type, shop=shop, created__gt=start_date, created__lt=end_date)
        else:
            messages.error(request, "Продаж в данный период не было")
            return redirect("sale_report_per_shop")
        # if Q(start_date) | Q(end_date):
        #     queryset_list = queryset_list.filter(created__range=[start_date, end_date])

       #============================Calculating Pay_Cards_Remainders per day======================= 
        product=Product.objects.get(imei='11111')    
        if RemainderHistory.objects.filter(shop=shop, imei=product.imei, created__lt=start_date).exists():
            rho_before=RemainderHistory.objects.filter(shop=shop, imei=product.imei, created__lt=start_date).latest("created")
            pay_card_remainder_start=rho_before.current_remainder
        else:
            pay_card_remainder_start=0
        if RemainderHistory.objects.filter(shop=shop, imei=product.imei, created__lt=end_date).exists():
            rho_latest=RemainderHistory.objects.filter(shop=shop, imei=product.imei, created__lt=end_date).latest("created")
            pay_card_remainder_current=rho_latest.current_remainder
        else:
            pay_card_remainder_current=0
       #===============================End of Calculating Pay Cards Remainder Module==================================

        #================Calculating Cash at the beginning of the day and at the end===========================
        if Cash.objects.filter(shop=shop, created__lt=start_date).exists():
            cho=Cash.objects.filter(shop=shop, created__lt=start_date).latest('created')
            cash_start = cho.current_remainder
        else:
            cash_start =0
        if Cash.objects.filter(shop=shop, created__lt=end_date).exists():
            cho_latest= Cash.objects.filter(shop=shop, created__lt=end_date).latest('created')
            cash_end=cho_latest.current_remainder
        #=====================Calculating Incoming Cash per day=============================
        cash_sum=0
        if Cash.objects.filter(shop=shop, created__gt=start_date, created__lt=end_date).exists():
            chos=Cash.objects.filter(shop=shop, created__gt=start_date, created__lt=end_date)
            for i in chos:
                cash_sum+=i.cash_in
        teko_sum=0
        if Cash.objects.filter(shop=shop, created__gt=start_date, created__lt=end_date, cho_type=teko_cash_in ).exists():
            teko_chos=Cash.objects.filter(shop=shop, created__gt=start_date, created__lt=end_date, cho_type=teko_cash_in)
            for i in teko_chos:
                teko_sum+=i.cash_in
        #=====================Calculating Outgoing Cash per day=============================
        total_expenses=0
        if Cash.objects.filter(shop=shop, created__gt=start_date, created__lt=end_date, cash_out__gt=0).exists():
            total_expenses_chos=Cash.objects.filter(shop=shop, created__gt=start_date, created__lt=end_date, cash_out__gt=0)
            for i in total_expenses_chos:
                total_expenses+=i.cash_out
        salary=0
        if Cash.objects.filter(shop=shop, created__gt=start_date, created__lt=end_date, cho_type=salary_cash_out).exists():
            salary_chos=Cash.objects.filter(shop=shop, created__gt=start_date, created__lt=end_date, cho_type=salary_cash_out)
            for i in salary_chos:
                salary+=i.cash_out
        expenses=0
        if Cash.objects.filter(shop=shop, created__gt=start_date, created__lt=end_date, cho_type=expenses_cash_out).exists():
            expenses_chos=Cash.objects.filter(shop=shop, created__gt=start_date, created__lt=end_date, cho_type=expenses_cash_out)
            for i in expenses_chos:
                expenses+=i.cash_out
        money_transfer=0
        if Cash.objects.filter(shop=shop, created__gt=start_date, created__lt=end_date, cho_type=money_transfer_cash_out).exists():
            money_transfer_chos=Cash.objects.filter(shop=shop, created__gt=start_date, created__lt=end_date, cho_type=money_transfer_cash_out)
            for i in money_transfer_chos:
                money_transfer+=i.cash_out
        money_return=0
        if Cash.objects.filter(shop=shop, created__gt=start_date, created__lt=end_date, cho_type=money_return_cash_out).exists():
            money_return_chos=Cash.objects.filter(shop=shop, created__gt=start_date, created__lt=end_date, cho_type=money_return_cash_out)
            for i in money_return_chos:
                money_return+=i.cash_out
   
    #===========================Calculaing Incoming Card Payments per day=====================
        card_sum=0
        if Card.objects.filter(shop=shop, created__gt=start_date, created__lt=end_date).exists():
            cards=Card.objects.filter(shop=shop, created__gt=start_date, created__lt=end_date)
            for i in cards:
                card_sum+=i.sum
    #====================Calculating Incoming Credit Payments per day=====================
        credit_sum=0
        if Credit.objects.filter(shop=shop, created__gt=start_date, created__lt=end_date).exists():
            credits=Credit.objects.filter(shop=shop, created__gt=start_date, created__lt=end_date)
            for i in credits:
                credit_sum+=i.sum
    #===========================Calculating cashback sum=================================
        cashback=0
        if Document.objects.filter(shop_sender=shop, created__gt=start_date, created__lt=end_date, posted=True).exists():
            docs=Document.objects.filter(shop_sender=shop, created__gt=start_date, created__lt=end_date, posted=True)
            for doc in docs:
                cashback+=doc.cashback_off

        array=[]
        for i in queryset:
            array.append(i.imei)
        array = set(array)#eliminating not unique imeis
        report_id = ReportTempId.objects.create()
        #calculating total quantity & sub_total per imei per day
        for i in array:
            queryset_list=queryset.filter(imei=i)
            quantity_out=0
            self_cost=0
            retail_sum=0
            for qs in queryset_list:
                quantity_out+=qs.outgoing_quantity
                self_cost+=qs.av_price*qs.outgoing_quantity
                retail_sum+=qs.sub_total
            product=Product.objects.get(imei=i)

            sale_rep = SaleReport.objects.create(
                report_id=report_id,
                product=product.name,
                category=product.category,
                quantity=quantity_out,
                av_sum=self_cost,
                retail_sum=retail_sum,
                margin=retail_sum - self_cost
            )
        #calculating total sales sum
        sale_report=SaleReport.objects.filter(report_id=report_id.id).order_by('category').order_by('product')
        total_sales=0
        for item in sale_report:
            total_sales+=item.retail_sum
     
        context = {
            "sale_report": sale_report,
            "shops": shops,
            "total_sales": total_sales,
            "shop": shop,
            "pay_card_remainder_start": pay_card_remainder_start,
            "pay_card_remainder_current": pay_card_remainder_current,
            "cash_sum": cash_sum,
            "teko_sum": teko_sum,
            "credit_sum": credit_sum,
            "card_sum": card_sum,
            "cash_start": cash_start,
            "cash_end": cash_end,
            "cashback": cashback,
            "money_return": money_return,
            "money_transfer": money_transfer,
            "expenses": expenses,
            "salary": salary,
            "total_expenses": total_expenses,
        }
        return render(request, "reports/sale_report_per_shop.html", context)
    else:
        context = {
            "shops": shops,
        }
        return render(request, "reports/sale_report_per_shop.html", context)

def sale_report_analytic(request):
    categories = ProductCategory.objects.all()
    products = Product.objects.all()
    shops = Shop.objects.all()
    suppliers = Supplier.objects.all()
    group_sales=Group.objects.get(name='sales')
    users = User.objects.filter(is_active=True, groups=group_sales ).order_by('username')
    if request.method == "POST":
        doc_type = DocumentType.objects.get(name="Продажа ТМЦ")
        queryset = RemainderHistory.objects.filter(rho_type=doc_type)
        sum = 0
        # category = request.POST["category"]
        category = request.POST.get("category", False)
        if category:
            category = ProductCategory.objects.get(id=category)
        shop = request.POST.get("shop", False)
        if shop:
            shop = Shop.objects.get(id=shop)
        supplier = request.POST.get("supplier", False)
        if supplier:
            supplier = Supplier.objects.get(id=supplier)
        user = request.POST.get("user", False)
        if user:
            user = User.objects.get(id=user)
        start_date = request.POST.get("start_date", False)
        if start_date:
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
        end_date = request.POST.get("end_date", False)
        if end_date:
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            end_date = end_date + timedelta(days=1)

        if category:
            #products=Product.objects.filter(category=category)
            queryset = queryset.filter(category=category)
        #else:
            #products=Product.objects.all()
        if shop:
            queryset = queryset.filter(shop=shop)
        # if supplier:
        #     queryset_list = queryset_list.filter(supplier=supplier)
        if user:
            queryset = queryset.filter(user=user)
            #user = user
        # if Q(start_date) | Q(end_date):
        #     queryset_list = queryset_list.filter(created__range=[start_date, end_date])
        if start_date:
            queryset = queryset.filter(created__gt=start_date)
        if end_date: 
            queryset = queryset.filter(created__lt=end_date)

        array=[]
        for i in queryset:
            array.append(i.imei)
        array = set(array)#eliminating not unique imeis
        #calculating payment types
        
        report_id = ReportTempId.objects.create()
        #calculating total quantity & sub_total per imei per day
        for i in array:
            queryset_list=queryset.filter(imei=i)
            quantity_out=0
            self_cost=0
            retail_sum=0
            for qs in queryset_list:
                quantity_out+=qs.outgoing_quantity
                self_cost+=qs.av_price*qs.outgoing_quantity
                retail_sum+=qs.sub_total
            product=Product.objects.get(imei=i)
            sale_rep = SaleReport.objects.create(
                report_id=report_id,
                product=product.name,
                imei=product.imei,
                quantity=quantity_out,
                av_sum=self_cost,
                retail_sum=retail_sum,
                margin=retail_sum - self_cost
            )
        #calculating total sales sum
        sale_report=SaleReport.objects.filter(report_id=report_id.id).order_by('product')
        total_sales=0
        av_sales=0
        profit=0
        for index, item in enumerate(sale_report):
            item.index = index
            item.save()
            total_sales+=item.retail_sum
            av_sales+=item.av_sum
            profit+=item.margin

        if shop:
            context = {
                "sale_report": sale_report,
                "categories": categories,
                "shops": shops,
                "suppliers": suppliers,
                "users": users,
                "total_sales": total_sales,
                "av_sales": av_sales,
                "profit": profit,
                "shop": shop,
                "report_id": report_id,
            }
            return render(request, "reports/sale_report_analytic.html", context)
        else:
            context = {
                "sale_report": sale_report,
                "categories": categories,
                "shops": shops,
                "suppliers": suppliers,
                "users": users,
                "total_sales": total_sales,
                "av_sales": av_sales,
                "profit": profit,
                #"shop": shop,
                "report_id": report_id,
                
            }
            return render(request, "reports/sale_report_analytic.html", context)
    else:
        context = {
            "categories": categories,
            "shops": shops,
            "suppliers": suppliers,
            "users": users,
        }
        return render(request, "reports/sale_report_analytic.html", context)

def effectiveness_report(request):
    products = Product.objects.all()
    shops = Shop.objects.all()
    group_sales=Group.objects.get(name='sales')
    users = User.objects.filter(is_active=True, groups=group_sales ).order_by('-username')
    if request.method == "POST":
        doc_type = DocumentType.objects.get(name="Продажа ТМЦ")
        shop = request.POST["shop"]
        shop=Shop.objects.get(id=shop)
        start_date = request.POST.get("start_date", False)
        start_date = request.POST["start_date"]
        start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
        end_date = request.POST["end_date"]
        end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
        end_date = end_date + timedelta(days=1)
        queryset = RemainderHistory.objects.filter(rho_type=doc_type, shop=shop, created__gt=start_date, created__lt=end_date)
        #creating a list of dates
        period= []
        while start_date <= end_date:
            period.append(start_date)
            start_date += timedelta(days=1)
        
        #print('======================')
        array=[]
        for date in period:
            for user in users:
                if queryset.filter(created__date=date, user=user).exists():
                    queryset_list=queryset.filter(created__date=date, user=user)
                    sum=0
                    for rho in queryset_list:
                        sum+=rho.retail_price
                    entry = [date, user, sum]
                    array.append(entry)
                

        # if Q(start_date) | Q(end_date):
        #     queryset_list = queryset_list.filter(created__range=[start_date, end_date])
        #user = request.POST.get("user", False)
        #if user:
        #   user = User.objects.get(id=user)
        #f user:
            #queryset = queryset.filter(user=user)
            #user = user
        
        
        #array = set(array)#eliminating not unique imeis
        
        report_id = ReportTempId.objects.create()
        #calculating total quantity & sub_total per imei per day
        for i in array:
            effectiveness_report = EffectivenessReport.objects.create(
                report_id=report_id,
                shop=shop,
                date=i[0],
                user=i[1],
                sum=i[2]
            )
                
        effectiveness_report=EffectivenessReport.objects.filter(report_id=report_id.id).order_by('date')
        total_sum = []
        for user in users: 
            if effectiveness_report.filter(user=user).exists():
                query=effectiveness_report.filter(user=user)
                user_sum=0
                counter_of_work_days = 0
                for i in query:
                    user_sum+=i.sum
                    counter_of_work_days+=1
                user_array=[user, user_sum, counter_of_work_days]
                total_sum.append(user_array)

        context = {
            "effectiveness_report": effectiveness_report,
            "shops": shops,
            #"users": users,
            "shop": shop,
            #"report_id": report_id,
            'total_sum': total_sum,
            
        }
        return render(request, "reports/effectiveness_report.html", context)
    else:
        context = {
            "shops": shops,
            "users": users,
        }
        return render(request, "reports/effectiveness_report.html", context)

def sale_report_excel (request, report_id):
    if request.user.is_authenticated:
        report_items=SaleReport.objects.filter(report_id=report_id)
        print(report_items.count())
        #==========================Convert to Excel module=========================================
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = (
            "attachment; filename=SaleRep_" + str(date) + ".xls"
        )

        # str(datetime.date.today())+'.xls'

        wb = xlwt.Workbook(encoding="utf-8")
        ws = wb.add_sheet('Period')

        # sheet header in the first row
        row_num = 0
        font_style = xlwt.XFStyle()
        columns = ['Модель', "IMEI", 'Кол-во', 'Опт', 'Розница', 'Прибыль']
        for col_num in range(len(columns)):
            ws.write(row_num, col_num + 1, columns[col_num], font_style)

        # sheet body, remaining rows
        font_style = xlwt.XFStyle()

        row_num = 1
        for item in report_items:
            col_num = 1
            ws.write(row_num, col_num, item.product, font_style)
            col_num +=1
            ws.write(row_num, col_num, item.imei, font_style)
            col_num +=1
            ws.write(row_num, col_num, item.quantity, font_style)
            col_num +=1
            ws.write(row_num, col_num, item.av_sum, font_style)
            col_num +=1
            ws.write(row_num, col_num, item.retail_sum, font_style)
            col_num +=1
            ws.write(row_num, col_num, item.margin, font_style)
            row_num +=1

        wb.save(response)
        return response

    else:
        logout(request)
        return redirect('login')

#==================================================================
#work for smartphones with unique IMEIs
def sale_report_per_supplier (request):
    if request.user.is_authenticated:
        suppliers=Supplier.objects.all()
        categories = ProductCategory.objects.all()
        products = Product.objects.all()
        doc_type = DocumentType.objects.get(name="Продажа ТМЦ")
        doc_type = DocumentType.objects.get(name="Продажа ТМЦ")
        doc_type_supply = DocumentType.objects.get(name="Поступление ТМЦ")
       
        if request.method == "POST":
            category = request.POST["category"]
            supplier = request.POST["supplier"]
            start_date = request.POST["start_date"]
            end_date = request.POST["end_date"]
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            end_date = end_date + timedelta(days=1)
            query = RemainderHistory.objects.filter(rho_type=doc_type, category=category,created__gt=start_date,  created__lt=end_date)
            arr ={}
            for item in query:
                if RemainderHistory.objects.filter(rho_type=doc_type_supply, category=category, imei=item.imei, supplier=supplier).exists():
                    number_of_docs=RemainderHistory.objects.filter(rho_type=doc_type_supply, category=category, imei=item.imei, supplier=supplier).count()
                    if number_of_docs < 2:#trying to eliminate 'returns' & next sales
                        product=RemainderHistory.objects.get(rho_type=doc_type_supply, category=category, imei=item.imei, supplier=supplier)
                        arr[item]=product

            #==========================Convert to Excel module=========================================
            response = HttpResponse(content_type="application/ms-excel")
            response["Content-Disposition"] = (
                "attachment; filename=SupplierRep_" + str(date) + ".xls"
            )

            # str(datetime.date.today())+'.xls'

            wb = xlwt.Workbook(encoding="utf-8")
            ws = wb.add_sheet('Period')

            # sheet header in the first row
            row_num = 0
            font_style = xlwt.XFStyle()
            columns = ['Дата продажи', "Модель", 'IMEI', 'Поставщик', "Дата поставки"]
            for col_num in range(len(columns)):
                ws.write(row_num, col_num + 1, columns[col_num], font_style)
            
            # sheet body, remaining rows
            font_style = xlwt.XFStyle()

            row_num = 1
            for key, value in arr.items():
                col_num = 1
                ws.write(row_num, col_num, str(key.created), font_style)
                col_num +=1
                ws.write(row_num, col_num, key.name, font_style)
                col_num +=1
                ws.write(row_num, col_num, key.imei, font_style)
                col_num +=1
                ws.write(row_num, col_num, value.supplier.name, font_style)
                col_num +=1
                ws.write(row_num, col_num, str(value.created), font_style)
                row_num +=1

            wb.save(response)
            return response
        else:
            context = {
                'suppliers': suppliers,
                'categories': categories,
                'products': products
            }
            return render (request, 'reports/sale_per_supplier.html' , context)
    else:
        logout(request)
        return redirect('login')

def delivery_report(request):
    if request.user.is_authenticated:
        suppliers = Supplier.objects.all()
        # deliveries = Delivery.objects.all()
        doc_type = DocumentType.objects.get(name="Поступление ТМЦ")
        queryset_list = RemainderHistory.objects.filter(rho_type=doc_type.id)
        if request.method == "POST":
            start_date = request.POST["start_date"]
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            end_date = request.POST["end_date"]
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            end_date = end_date + timedelta(days=1)
            queryset_list = queryset_list.filter(created__gte=start_date, created__lte=end_date)
            report_id = ReportTempId.objects.create()
            for supplier in suppliers:
                counter=0
                for item in queryset_list:
                    if item.supplier == supplier:
                        counter+=item.sub_total
                if counter > 0:              
                    report_item=DeliveryReport.objects.create(
                        report_id=report_id,
                        supplier=supplier.name,
                        sum=counter
                    )
            
            query=DeliveryReport.objects.filter(report_id=report_id)

            #==========================Convert to Excel module=========================================
            response = HttpResponse(content_type="application/ms-excel")
            response["Content-Disposition"] = (
                "attachment; filename=DeliveryRep_"+ str(end_date) + ".xls"
            )

            # str(datetime.date.today())+'.xls'

            wb = xlwt.Workbook(encoding="utf-8")
            ws = wb.add_sheet('Delivery')

            # sheet header in the first row
            row_num = 0
            font_style = xlwt.XFStyle()
            columns = ["Поставщик", "Сумма"]
            for col_num in range(len(columns)):
                ws.write(row_num, col_num + 1, columns[col_num], font_style)
            
            # sheet body, remaining rows
            font_style = xlwt.XFStyle()

            row_num = 1
            for item in query:
                col_num = 1
                ws.write(row_num, col_num, item.supplier, font_style)
                col_num +=1
                ws.write(row_num, col_num, item.sum, font_style)
                row_num +=1

            wb.save(response)
            return response
        #================================End of Excel Module========================================




            # qs=query.values('supplier', 'sum',)
            # data=pd.DataFrame.from_records(qs)
            # data.to_excel('delivery.xlsx')
           
            return render(request, "reports/delivery_report.html")
        
        else:
            return render(request, "reports/delivery_report.html")
    else:
        logout(request)
        return redirect('login')

def delivery_report_per_supplier(request):
    categories = ProductCategory.objects.all()
    suppliers = Supplier.objects.all()
    # deliveries = Delivery.objects.all()
    doc_type = DocumentType.objects.get(name="Поступление ТМЦ")
    queryset_list = RemainderHistory.objects.filter(rho_type=doc_type.id)
    if request.method == "POST":
        # shop = request.POST["shop"]
        category = request.POST["category"]
        supplier = request.POST["supplier"]
        # user = request.POST["user"]
        start_date = request.POST["start_date"]
        end_date = request.POST["end_date"]
        if supplier:
            queryset_list = queryset_list.filter(supplier=supplier)
        if start_date:
            queryset_list = queryset_list.filter(created__gte=start_date)
        if end_date:
            queryset_list = queryset_list.filter(created__lte=end_date)
        if category:
            queryset_list = queryset_list.filter(category=category)
        context = {
            "categories": categories,
            "suppliers": suppliers,
            "queryset_list": queryset_list,
            "sum": sum,
        }
        return render(request, "reports/delivery_report.html", context)
    else:
        context = {
            "categories": categories,
            "suppliers": suppliers,
            "queryset_list": queryset_list,
        }
    return render(request, "reports/delivery_report.html", context)
#==========================================================
def expenses_report(request):
    if request.user.is_authenticated:
        #users=User.objects.all()
        doc_type = DocumentType.objects.get(name="РКО (хоз.расходы)")
        shops = Shop.objects.filter(active=True, retail=True).order_by("name")
        queryset_list = Cash.objects.filter(cho_type=doc_type.id)
        if request.method == "POST":
            start_date = request.POST["start_date"]
            # converting HTML date format (2021-07-08T01:05) to django format (2021-07-10 01:05:00)
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            end_date = request.POST["end_date"]
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            end_date = end_date + timedelta(days=1)
            queryset_list = queryset_list.filter(created__gte=start_date, created__lte=end_date)
            report_id = ReportTempId.objects.create()
            for shop in shops:
                counter=0
                for item in queryset_list:
                    if item.shop == shop:
                        counter+=item.cash_out               
                report_item=ExpensesReport.objects.create(
                    report_id=report_id,
                    shop=shop.name,
                    sum=counter,
                )
            
            query=ExpensesReport.objects.filter(report_id=report_id)

            #==========================Convert to Excel module=========================================
            response = HttpResponse(content_type="application/ms-excel")
            response["Content-Disposition"] = (
                "attachment; filename=ExpenseRep_"+ str(end_date) + ".xls"
            )

            # str(datetime.date.today())+'.xls'

            wb = xlwt.Workbook(encoding="utf-8")
            ws = wb.add_sheet('Salary')

            # sheet header in the first row
            row_num = 0
            font_style = xlwt.XFStyle()
            columns = ["Точка", "Сумма"]
            for col_num in range(len(columns)):
                ws.write(row_num, col_num + 1, columns[col_num], font_style)
            
            # sheet body, remaining rows
            font_style = xlwt.XFStyle()

            row_num = 1
            for item in query:
                col_num = 1
                ws.write(row_num, col_num, item.shop, font_style)
                col_num +=1
                ws.write(row_num, col_num, item.sum, font_style)
                row_num +=1

            wb.save(response)
            return response
        #================================End of Excel Module========================================

            # qs=query.values('shop', 'sum',)
            # data=pd.DataFrame.from_records(qs)
            # data.to_excel('expenses.xlsx')
            # os.system('start excel.exe expenses.xlsx')
            # return render(request, "reports/expenses_report.html")
        
        else:
            return render(request, "reports/expenses_report.html")
    else:
        logout(request)
        return redirect('login')

def salary_report(request):
    if request.user.is_authenticated:
        group_sales=Group.objects.get(name='sales')
        #users = User.objects.filter(is_active=True, groups=group_sales ).order_by('last_name')
        users = User.objects.filter(groups=group_sales ).order_by('last_name')
        doc_type = DocumentType.objects.get(name="РКО (зарплата)")
        # shops = Shop.objects.filter(active=True, retail=True).order_by("name")
        queryset_list = Cash.objects.filter(cho_type=doc_type.id)
        if request.method == "POST":
            start_date = request.POST["start_date"]
            end_date = request.POST["end_date"]
            # converting HTML date format (2021-07-08T01:05) to django format (2021-07-10 01:05:00)
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            tdelta = datetime.timedelta(days=1)
            end_date = end_date + tdelta
            queryset_list = queryset_list.filter(created__gte=start_date, created__lte=end_date)
            report_id = ReportTempId.objects.create()
            for user in users:
                counter=0
                for item in queryset_list:
                    if item.cash_receiver == user:
                        counter+=item.cash_out

                if counter>0:
                    report_item=SalaryReport.objects.create(
                        report_id=report_id,
                        user=user.last_name,
                        sum=counter,
                    )
            
            query=SalaryReport.objects.filter(report_id=report_id)

            #==========================Convert to Excel module=========================================
            response = HttpResponse(content_type="application/ms-excel")
            response["Content-Disposition"] = (
                "attachment; filename=SalaryRep_"+ str(end_date) + ".xls"
            )

            # str(datetime.date.today())+'.xls'

            wb = xlwt.Workbook(encoding="utf-8")
            ws = wb.add_sheet('Salary')

            # sheet header in the first row
            row_num = 0
            font_style = xlwt.XFStyle()
            columns = ["ФИО", "Сумма"]
            for col_num in range(len(columns)):
                ws.write(row_num, col_num + 1, columns[col_num], font_style)
            
            # sheet body, remaining rows
            font_style = xlwt.XFStyle()

            row_num = 1
            for item in query:
                col_num = 1
                ws.write(row_num, col_num, item.user, font_style)
                col_num +=1
                ws.write(row_num, col_num, item.sum, font_style)
                row_num +=1

            wb.save(response)
            return response
        #================================End of Excel Module========================================

            # qs=query.values('user', 'sum',)
            # data=pd.DataFrame.from_records(qs)
            
            # data.to_excel('app_reports/salary.xlsx')
            # os.system('start excel.exe salary.xlsx')
            # return render(request, "reports/salary_report.html")
        else:
            # context = {
            #     'users': users
            # }
            return render(request, "reports/salary_report.html")
    else:
        logout(request)
        return redirect('login')
# =======================================================
def remainder_report(request):
    if request.user.is_authenticated:
        users = Group.objects.get(name="sales").user_set.all()
        group = Group.objects.get(name="admin").user_set.all()
        categories = ProductCategory.objects.all()
        shops = Shop.objects.all()
        if request.method == "POST":
            category = request.POST["category"]
            category = ProductCategory.objects.get(id=category)
            if request.user in group:
                shop = request.POST["shop"]
                shop = Shop.objects.get(id=shop)
            else:
                session_shop = request.session["session_shop"]
                shop = Shop.objects.get(id=session_shop)
            # ==============Time Module=========================================
            date = request.POST.get("date", False)
            if date:
                date = datetime.datetime.strptime(date, "%Y-%m-%d")
            else:
                date = datetime.date.today()
            tdelta_1 = datetime.timedelta(days=1)
            date = date + tdelta_1
            return redirect("remainder_report_output", shop.id, category.id, date)
        else:
            context = {
                "shops": shops,
                "categories": categories,
            }
            return render(request, "reports/remainder_report.html", context)
    else:
        auth.logout(request)
        return redirect("login")

def remainder_report_ver_1(request):
    if request.user.is_authenticated:
        users = Group.objects.get(name="sales").user_set.all()
        group = Group.objects.get(name="admin").user_set.all()
        categories = ProductCategory.objects.all()
        shops = Shop.objects.all()
        if request.method == "POST":
            category = request.POST["category"]
            category = ProductCategory.objects.get(id=category)
            if request.user in group:
                shop = request.POST["shop"]
                shop = Shop.objects.get(id=shop)
            else:
                session_shop = request.session["session_shop"]
                shop = Shop.objects.get(id=session_shop)

            return redirect("remainder_report_output_ver_1", shop.id, category.id)
        else:
            context = {
                "shops": shops,
                "categories": categories,
            }
            return render(request, "reports/remainder_report_ver_1.html", context)

    else:
        auth.logout(request)
        return redirect("login")

def remainder_report_output_ver_1(request, shop_id, category_id):
    if request.user.is_authenticated:
        shop = Shop.objects.get(id=shop_id)
        category = ProductCategory.objects.get(id=category_id)
        current_remainders=RemainderCurrent.objects.filter(category=category, shop=shop)
        for item in current_remainders:
            print(item.name)
        context = {
        
            "shop": shop, 
            "category": category,
            "current_remainders": current_remainders,
            }
        return render(request, "reports/remainder_report_output_ver_1.html", context)
    else:
        auth.logout(request)
        return redirect("login")

def remainder_report_excel(request, shop_id, category_id, date):
    users = Group.objects.get(name="sales").user_set.all()
    if request.user.is_authenticated:
        date = date
        shop = Shop.objects.get(id=shop_id)
        category = ProductCategory.objects.get(id=category_id)
        products = Product.objects.filter(category=category)
        report_id = ReportTempId.objects.create()
        for product in products:
            imei = product.imei
            if RemainderHistory.objects.filter(shop=shop, imei=imei, created__lte=date).exists():
                rho = RemainderHistory.objects.filter(shop=shop, imei=imei, created__lte=date).latest("created")
                if rho.current_remainder > 0:
                    if shop.retail == True:
                        price = rho.retail_price
                    elif shop.subdealer == True:
                        price=rho.retail_price
                    else:
                        price = rho.wholesale_price
                    report = ReportTemp.objects.create(
                        report_id=report_id,
                        name=rho.name,
                        imei=rho.imei,
                        end_remainder=rho.current_remainder,
                        price=price,
                    )
        # report_query=ReportTemp.objects.filter(report_id=report_id)
        # query = report_query.values("name", "imei", "end_remainder", "price")
        # data=pd.DataFrame.from_records(query)
        # data.to_excel(f'RemainderReport_{date}.xlsx', index=False)
        # data.to_excel('RemainderReport.xlsx', index=False)

#=======================Uploading to Excel Module===================================
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = (
            "attachment; filename=Remainder_" + str(datetime.date.today()) + ".xls"
        )

        wb = xlwt.Workbook(encoding="utf-8")
        ws = wb.add_sheet("Remainder")

        # sheet header in the first row
        row_num = 0
        font_style = xlwt.XFStyle()

        columns = ["Name", "IMEI", "Quantity", "Price"]
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)

        # sheet body, remaining rows
        font_style = xlwt.XFStyle()
        report_query = ReportTemp.objects.filter(report_id=report_id)
        query = report_query.values_list("name", "imei", "end_remainder", "price")

        for row in query:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, str(row[col_num]), font_style)
        wb.save(response)
        return response
#=======================End of Excel Upload Module================================

        for i in report_query:
            i.delete()
        report_id.delete()
        # if request.user in users:
        #     return redirect ('sale_interface')
        # else:
        #     return redirect("log")
    else:
        auth.logout(request)
        return redirect("login")

def remainder_report_output(request, shop_id, category_id, date):
    if request.user.is_authenticated:
        date = date
        shop = Shop.objects.get(id=shop_id)
        category = ProductCategory.objects.get(id=category_id)
        array = []
        products = Product.objects.filter(category=category).order_by(
            "name"
        )  # order_by name lets us created an array sorted in alphabeticatl order for further processing as a table
        messages.success(request, 'Для содания документа "Переоценка ТМЦ" выделите необходимые позиции и нажмите кнопку в конце страницы')
        for product in products:
            imei = product.imei
            if RemainderHistory.objects.filter(shop=shop, imei=imei, created__lte=date).exists():
                rho = RemainderHistory.objects.filter(
                    shop=shop, imei=imei, created__lte=date).latest("created")
                if rho.current_remainder > 0:
                    array.append(rho)
        arr_length=len(array)
        for arr, i in zip(array, range(arr_length)):
            arr.number = i + 1
            arr.save()
        context = {
            "date": date, 
            "shop": shop, 
            "array": array, 
            #"arr_length": arr_length,
            "category": category
            }
        return render(request, "reports/remainder_report_output.html", context)
    else:
        auth.logout(request)
        return redirect("login")

def remainder_report_dynamic(request):
    products = Product.objects.all()
    categories = ProductCategory.objects.all()
    shops = Shop.objects.all()
    if request.method == "POST":
        report_id = ReportTempId.objects.create()
        #date indicates 00.00 (the beginning of the date indicated)
        date_start = request.POST["date_start"]
        date_start = datetime.datetime.strptime(date_start, "%Y-%m-%d")
        date_end = request.POST["date_end"]
        date_end = datetime.datetime.strptime(date_end, "%Y-%m-%d")
        date_end = date_end + timedelta(days=1)
        category = request.POST["category"]
        shop = request.POST["shop"]
        shop = Shop.objects.get(id=shop)
        products = Product.objects.filter(category=category)
        i=0
        for product in products:
            #checking rhos before the period indicated
            if RemainderHistory.objects.filter(shop=shop, imei=product.imei, created__lt=date_start).exists():
                rho_before_latest=RemainderHistory.objects.filter(shop=shop, imei=product.imei, created__lt=date_start).latest('created')
                if rho_before_latest.current_remainder > 0:
                    remainder_start=rho_before_latest.current_remainder
                else:
                    remainder_start=0
            else:
                remainder_start=0
            #checking rhos during the period indicated
            if RemainderHistory.objects.filter(shop=shop, imei=product.imei, created__gt=date_start, created__lt=date_end).exists():
                queryset=RemainderHistory.objects.filter(shop=shop, imei=product.imei, created__gt=date_start, created__lt=date_end)
                q_in = 0
                q_out = 0
                for qs in queryset:
                    q_in += qs.incoming_quantity
                    q_out += qs.outgoing_quantity
                remainder_end = remainder_start - q_out + q_in
            else:
                remainder_end=remainder_start
                q_in=0
                q_out=0
            if remainder_start == 0 and q_in==0 and q_out==0:
                pass
            else: 
                report = ReportTemp.objects.create(
                    report_id=report_id,
                    name=product.name,
                    imei=product.imei,
                    quantity_in=q_in,
                    quantity_out=q_out,
                    initial_remainder=remainder_start,
                    end_remainder=remainder_end,
                )        
        reports = ReportTemp.objects.filter(report_id=report_id).order_by('name')
        context = {
            "reports": reports,
            "shops": shops,
            "categories": categories,
            "date_start": date_start,
            "date_end": date_end,
            "shop": shop,
        }
        return render(request, "reports/remainder_report_dynamic.html", context)
    context = {
        "shops": shops,
        "categories": categories,
    }
    return render(request, "reports/remainder_report_dynamic.html", context)

def remainder_general_report (request):
    if request.user.is_authenticated:
        transfer= DocumentType.objects.get(name='Перемещение ТМЦ')
        users = Group.objects.get(name="sales").user_set.all()
        group = Group.objects.get(name="admin").user_set.all()
        categories = ProductCategory.objects.all()
        shops = Shop.objects.all()
        if request.method == "POST":
            if request.user in group:
                category = request.POST["category"]
                category = ProductCategory.objects.get(id=category)
                # ==============Time Module=========================================
                date = request.POST["date"]
                date = datetime.datetime.strptime(date, "%Y-%m-%d")
                tdelta_1 = datetime.timedelta(days=1)
                date = date + tdelta_1
                products=Product.objects.filter(category=category)
                arr =[]
                for product in products:
                    imei=product.imei
                    if RemainderHistory.objects.filter(imei=imei, created__lte=date).exists():
                        rho = RemainderHistory.objects.filter(imei=imei, created__lte=date).latest("created")
                        if rho.rho_type == transfer: #перемещение создает два rho с одинаковым временем и нам нужно их разделить.
                            document=rho.document
                            shop_receiver=document.shop_receiver
                            try:
                                rho=RemainderHistory.objects.get(document=document, imei = imei, shop=shop_receiver)
                            except:
                                string=f'{imei} не уникален в документе #{document} от {document.created}.'
                                messages.error(request,  string)
                                return redirect("remainder_general_report")
                           
                        if rho.current_remainder > 0:
                            arr.append(rho)
        #==========================Convert to Excel module=========================================
                response = HttpResponse(content_type="application/ms-excel")
                response["Content-Disposition"] = (
                    "attachment; filename=RemainderRep_" + str(date) + ".xls"
                )

                # str(datetime.date.today())+'.xls'

                wb = xlwt.Workbook(encoding="utf-8")
                ws = wb.add_sheet('Period')

                # sheet header in the first row
                row_num = 0
                font_style = xlwt.XFStyle()
                columns = ["Модель", 'IMEI', 'Кол-во', 'Опт', 'Розница']
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num + 1, columns[col_num], font_style)
                
                # sheet body, remaining rows
                font_style = xlwt.XFStyle()

                row_num = 1
                for item in arr:
                    col_num = 1
                    ws.write(row_num, col_num, item.name, font_style)
                    col_num +=1
                    ws.write(row_num, col_num, item.imei, font_style)
                    col_num +=1
                    ws.write(row_num, col_num, item.current_remainder, font_style)
                    col_num +=1
                    ws.write(row_num, col_num, item.wholesale_price, font_style)
                    col_num +=1
                    ws.write(row_num, col_num, item.retail_price, font_style)
                    row_num +=1

                wb.save(response)
                return response
        #================================End of Excel Module========================================
            else:
                logout(request)
                return redirect('login')
        else:
            context = {
                'categories': categories
            }
            return render (request, 'reports/remainder_general_report.html', context)


    else:
        logout(request)
        return redirect('login')

def remainder_list(request, shop_id, category_id):
    shop = Shop.objects.get(id=shop_id)
    category = ProductCategory.objects.get(id=category_id)
    # remainders=RemainderCurrent.objects.filter(shop=shop, category=category).order_by('name').values_list()
    # remainders=RemainderCurrent.objects.filter(shop=shop, category=category).order_by('name')
    remainders = (
        RemainderCurrent.objects.filter(shop=shop, category=category)
        .order_by("name")
        .values()
    )
    print("##########################")
    print(remainders)
    print("##########################")
    df = pd.DataFrame(remainders)
    print(df.shape)  # displays the number of rows & columns
    print("##########################")
    print(df)
    print("##########################")
    # df.to_excel('data.xlsx')
    context = {
        #'remainders':remainders,
        "df": df.to_html(),
        "category": category,
    }
    return render(request, "reports/remainder_report_output.html", context)

# ======================================================================
def item_report(request):
    if request.user.is_authenticated:
        if request.method == "POST":
            imei = request.POST["imei"]
            start_date = request.POST["start_date"]
            end_date = request.POST["end_date"]
            imei = request.POST["imei"]
            if '/' in imei:
                imei=imei.replace('/', '_')
            imei=imei.replace(" ","")#getting rid of spaces.
            try:
                product=Product.objects.get(imei=imei)
                queryset_list = RemainderHistory.objects.filter(imei=imei).order_by("created")
                if start_date:
                    queryset_list = queryset_list.filter(created__gte=start_date).order_by('created')
                if end_date:
                    # converting HTML date format (2021-07-08T01:05) to django format (2021-07-10 01:05:00)
                    end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
                    # adding time delta to cover docs created after year:month:day:00:00 till the end of the day 23:59
                    tdelta = datetime.timedelta(days=1)
                    end_date = end_date + tdelta
                    queryset_list = queryset_list.filter(created__lte=end_date).order_by('created')

                context = {
                    "queryset_list": queryset_list,
                    "product": product,
                }
                return render(request, "reports/item_report.html", context)
            except:
                messages.error(request, "Наименование с таким IMEI не содержится в базе данных.")
                return redirect("item_report")
        else:
            return render(request, "reports/item_report.html")
    else:
        auth.logout(request)
        return redirect("login")

# ====================================================================
def payment_report (request):
    if request.user.is_authenticated:
        shops=Shop.objects.all()
        if request.method=="POST":
            start_date = request.POST["start_date"]
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            end_date = request.POST["end_date"]
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            end_date = end_date + timedelta(days=1)
            cash=Cash.objects.filter(created__gt=start_date, created__lt=end_date)
            card=Card.objects.filter(created__gt=start_date, created__lt=end_date)
            credit=Credit.objects.filter(created__gt=start_date, created__lt=end_date)
            shop = request.POST.get("shop", False)
            if shop:
                shop = Shop.objects.get(id=shop)
                cash = cash.filter(shop=shop)
                card = card.filter(shop=shop)
                credit=credit.filter(shop=shop)
            cash_sum=0
            for item in cash:
                cash_sum+=item.cash_in
            card_sum=0
            for item in card:
                card_sum+=item.sum
            credit_sum=0
            for item in credit:
                credit_sum+=credit.sum
            context = {
                "cash_sum": cash_sum,
                "card_sum": card_sum,
                "credit_sum": credit_sum,
                "shops": shops
            }
            return render(request, "reports/payment_report.html", context)
        else:
            context = {
                'shops': shops
            }
            return render(request, "reports/payment_report.html", context)
    else:
        auth.logout(request)
        return redirect("login")

def daily_shop_rep(request):
    if request.user.is_authenticated:
        session_shop=request.session['session_shop']
        shop=Shop.objects.get(id=session_shop)
        rhos=RemainderHistory.objects.filter(created__date=datetime.today, shop=shop)
        
        pass
    else:
        auth.logout(request)
        return redirect("login")

def cash_report(request):
    if request.user.is_authenticated:
        session_shop=request.session['session_shop']
        shop=Shop.objects.get(id=session_shop)
        if request.method=='POST':
            # start_date = request.POST.get("start_date", False)
            start_date = request.POST["start_date"]
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            end_date = request.POST["end_date"]
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            end_date = end_date + timedelta (days=1)
               
       
            queryset_list = Cash.objects.filter(shop=shop, created__gt=start_date, created__lt=end_date)
            context = {
                "queryset_list": queryset_list
                }
            return render(request, 'reports/cash_report.html', context)
        else:
            return render(request, "reports/cash_report.html")
    else:
        auth.logout(request)
        return redirect("login")

def credit_report(request):
    shops = Shop.objects.all()
    credits = Credit.objects.all()
    users = User.objects.all()
    if request.method == "POST":
        start_date = request.POST["start_date"]
        start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
        end_date = request.POST["end_date"]
        end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
        end_date = end_date + timedelta(days=1)
        shop = request.POST.get("shop", False)
        if shop:
            shop = Shop.objects.get(id=shop)
        user = request.POST.get("user", False)
        if user:
            user = User.objects.get(id=user)
        credit_report = Credit.objects.filter(
            created__gte=start_date, created__lte=end_date
        )
        if shop:
            credit_report = credit_report.filter(shop=shop)
        if user:
            credit_report = credit_report.filter(user=user)
        context = {"shops": shops, "users": users, "credit_report": credit_report}
        return render(request, "reports/credit_report.html", context)

    context = {"shops": shops, "users": users}
    return render(request, "reports/credit_report.html", context)


    if request.user.is_authenticated:
        shops = Shop.objects.all()
        cards = Card.objects.all()
        if request.method == "POST":
            start_date = request.POST["start_date"]
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            end_date = request.POST["end_date"]
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            end_date = end_date + timedelta(days=1)
            shop = request.POST.get("shop", False)
            if shop:
                shop = Shop.objects.get(id=shop)
    
            card_report = Card.objects.filter(
                created__gte=start_date, created__lte=end_date
            )
            if shop:
                card_report = card_report.filter(shop=shop)
                total_sum=0
                for item in card_report:
                    total_sum+=item.sum
            context = {
                "shops": shops,
                "card_report": card_report,
                "total_sum": total_sum
                }
            return render(request, "reports/card_report.html", context)
        else:
            context = {
                "shops": shops, 
            }
            return render(request, "reports/card_report.html", context)
    else:
        auth.logout(request)
        return redirect("login")

#====================Pay_Card_Reports==============================
def card_report(request):
    if request.user.is_authenticated:
        shops = Shop.objects.filter(active=True, retail=True)
        cards = Card.objects.all()
        if request.method == "POST":
            start_date = request.POST["start_date"]
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            end_date = request.POST["end_date"]
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            end_date = end_date + timedelta(days=1)
            file = request.FILES["file_name"]

            report_id = ReportTempId.objects.create()

            df1 = pd.read_excel(file)
            cycle = len(df1)
            #processing data from Card reports from shops
            our_card_report=Card.objects.filter(created__gte=start_date, created__lte=end_date)
            for shop in shops:
                total_sum_retail=0
                our_card_report_per_shop=our_card_report.filter(shop=shop)
                for i in our_card_report_per_shop:
                    total_sum_retail+=int(i.sum)
                
                item= AcquiringReport.objects.create (
                    report_id=report_id,
                    sum_retail=total_sum_retail,
                    TID=shop.TID,
                    shop=str(shop.name)
                )
            
                for i in range(cycle):
                    row = df1.iloc[i]#reads each row of the df1 one by one
                    description=row.TID
                    sum=row.Sum
                    if str(shop.TID) in description:
                        item.sum_bank+=int(row.Sum)
                        item.save()
    
            #qs=AcquiringReport.objects.filter(report_id=report_id).values_list()
            report=AcquiringReport.objects.filter(report_id=report_id)
            qs=report.values('TID', 'shop', 'sum_bank', 'sum_retail')
            data=pd.DataFrame.from_records(qs)
            data.to_excel('data.xlsx')

            context = {
                "shops": shops, 
            }
            return render(request, "reports/card_report.html", context)

            # #==========================Convert to Excel module=========================================
            # response = HttpResponse(content_type="application/ms-excel")
            # response["Content-Disposition"] = (
            #     "attachment; filename=BonusRep_" + str(date) + ".xls"
            # )
            # # str(datetime.date.today())+'.xls'

            # wb = xlwt.Workbook(encoding="utf-8")
            # ws = wb.add_sheet('Сверка')

            # # sheet header in the first row
            # row_num = 0
            # col_num = 0
            # font_style = xlwt.XFStyle()
            # columns = ['TID','Сумма_банк', "Сумма_Ритейл", ]
            # # for category in categories:
            # #     columns.append(category.name)
            # # columns.append('Кредиты %')
            # # columns.append('Тяжелые тарифы (100)')
            # for col_num in range(len(columns)):
            #     #ws.write(row_num, col_num + 1, columns[col_num], font_style)
            #     ws.write(row_num, col_num, columns[col_num], font_style)

            # sheet body, remaining rows
            # font_style = xlwt.XFStyle()

            # row_num = 1
            #for item in monthly_report:
            #for i, c, x, y in zip(dict_bank.items(), dict_our.items()):
            #for i in dict_bank:
            # for i in range(cycle)
            #     col_num=0
            #     ws.write(row_num, col_num, key, font_style)
            #     col_num += 1
            #     ws.write(row_num, col_num, dict_bank[key], font_style)
            #     col_num += 1
            #     ws.write(row_num, col_num, dict_our[key1], font_style)
            #     row_num += 1
                
            
            # query_set = MonthlyBonus.objects.filter().values()
            # data = pd.DataFrame(query_set)
            # data = data.drop("id", 1)
            # data = data.set_index("user_name")
            # data.to_excel("D:/Аналитика/Фин_отчет/Текущие/2021/data.xlsx")

            # monthly_bonus_reports = MonthlyBonus.objects.all()
            # for i in monthly_bonus_reports:
            #     i.delete()

            # monthly_bonus_reports = AcquiringReport.objects.all()
            # for i in monthly_bonus_reports:
            #     i.delete()

            # wb.save(response)
            # return response


            


        #     shop = request.POST.get("shop", False)
        #     if shop:
        #         shop = Shop.objects.get(id=shop)
    
        #     card_report = Card.objects.filter(created__gte=start_date, created__lte=end_date)
        #     if shop:
        #         card_report = card_report.filter(shop=shop)
        #         total_sum=0
        #         for item in card_report:
        #             total_sum+=item.sum
        #     context = {
        #         "shops": shops,
        #         "card_report": card_report,
        #         "total_sum": total_sum
        #         }
        #     return render(request, "reports/card_report.html", context)
        else:
            context = {
                "shops": shops, 
            }
            return render(request, "reports/card_report.html", context)
    

    else:
        auth.logout(request)
        return redirect("login")

def daily_pay_card_rep_per_shop (request):
    if request.user.is_authenticated:
        shops=Shop.objects.all()
        if request.method == "POST":
            shop = request.POST.get("shop", False)
            if shop:
                shop=Shop.objects.get(shop=shop)
            else:
                session_shop=request.session['session_shop']
                shop=Shop.objects.get(id=session_shop)
            date = request.POST.get("date", False)
            if date:
                # converting HTML date format (2021-07-08T01:05) to django format (2021-07-10 01:05:00)
                date = datetime.datetime.strptime(date, "%Y-%m-%d")
            else:
                date = datetime.date.today()
            product=Product.objects.get(imei='11111')
            if RemainderHistory.objects.filter(shop=shop, imei=product.imei, created__date=date).exists():
                rhos=RemainderHistory.objects.filter(shop=shop, imei=product.imei, created__date=date)
            else:
                messages.error(request,  'Приход и расхода КЭО не было')
                return redirect("daily_pay_card_rep_per_shop")
            context = {
                "rhos": rhos,
                "shops": shops
            }
            return render(request, "reports/daily_pay_card_rep.html", context)
        else:
            context = {
                "shops": shops
            }
            return render(request, "reports/daily_pay_card_rep.html", context)

    else:
        auth.logout(request)
        return redirect("login")

def daily_pay_card_rep_general(request):
    shops = Shop.objects.all()
    if request.method == "POST":
        start_date = request.POST["start_date"]
        start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
        end_date = request.POST["end_date"]
        end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
        end_date = end_date + timedelta(days=1)
        #product = Product.objects.get(imei=11111)
        items=['11111', '99999']
        report_id = ReportTempId.objects.create()
        doc_type=DocumentType.objects.get(name='Переоценка ТМЦ')
        # ========================================================
        for shop in shops:
            for item in items:
                product=Product.objects.get(imei=item)
                if RemainderHistory.objects.filter(imei=product.imei, created__lt=start_date, shop=shop).exists():
                    rho_latest = RemainderHistory.objects.filter(imei=product.imei, created__lt=start_date, shop=shop).latest('created')
                    pre_remainder=rho_latest.current_remainder
                else:
                    pre_remainder=0
                if RemainderHistory.objects.filter(imei=product.imei, created__lt=end_date, shop=shop).exists():
                    rho_last = RemainderHistory.objects.filter(imei=product.imei, created__lt=end_date, shop=shop).latest('created')
                    current_remainder=rho_last.current_remainder
                else:
                    current_remainder=0
                if RemainderHistory.objects.filter(imei=product.imei, shop=shop, created__gt=start_date, created__lt=end_date).exclude(rho_type=doc_type).exists():
                    rhos = RemainderHistory.objects.filter(imei=product.imei, shop=shop, created__gt=start_date, created__lt=end_date).exclude(rho_type=doc_type)
                    incoming_quantity = 0
                    outgoing_quantity = 0
                    for rho in rhos:
                        incoming_quantity += rho.incoming_quantity
                        outgoing_quantity += rho.outgoing_quantity
                else:
                    incoming_quantity = 0
                    outgoing_quantity = 0
            # =========================================================
                pay_card_rep = PayCardReport.objects.create(
                    report_id=report_id,
                    shop=shop,
                    product=product,
                    pre_remainder=pre_remainder,
                    incoming_quantity=incoming_quantity,
                    outgoing_quantity=outgoing_quantity,
                    current_remainder=current_remainder,
                )
        #==========================Convert to Excel module=========================================
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = (
            "attachment; filename=DailRep_" + str(date) + ".xls"
        )
        # str(datetime.date.today())+'.xls'

        wb = xlwt.Workbook(encoding="utf-8")
        ws = wb.add_sheet('PayCardReport')



        # ======================sheet header in the first row
        row_num = 0
        font_style = xlwt.XFStyle()
        columns = []
        for shop in shops:
            columns.append(shop.name)
        for col_num in range(len(columns)):
            ws.write(row_num, col_num + 1, columns[col_num], font_style)
        #===========================================================

        # sheet body, remaining rows
        font_style = xlwt.XFStyle()
        daily_rep = PayCardReport.objects.filter(report_id=report_id)


        
        #Creating table for pay_cards_100 =================================
        col_num = 1
        for shop in shops:
            product=Product.objects.get(imei="11111")
            query_list = daily_rep.get(shop=shop.name, product=product)
            row_num = 1
            ws.write(row_num, col_num, query_list.pre_remainder, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.incoming_quantity, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.outgoing_quantity, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.current_remainder, font_style)
            col_num += 1

        #=========Creating first column========================
        product_list = [
            '100_остаток на утро',
            '100_приход',
            '100_расход',
            '100_остаток на вечер',
        ]
        row_num = 1
        for i in product_list:
            ws.write(row_num, 0, i, font_style)
            row_num = row_num + 1
        #======================================================


        #Creating table for pay_cards_500 =================================
        col_num = 1
        for shop in shops:
            product=Product.objects.get(imei="99999")
            query_list = daily_rep.get(shop=shop.name, product=product)
            row_num = 6
            ws.write(row_num, col_num, query_list.pre_remainder, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.incoming_quantity, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.outgoing_quantity, font_style)
            row_num += 1
            ws.write(row_num, col_num, query_list.current_remainder, font_style)
            col_num += 1

        #=========Creating first column========================
        product_list = [
            '500_остаток на утро',
            '500_приход',
            '500_расход',
            '500_остаток на вечер',
        ]
        row_num = 6
        for i in product_list:
            ws.write(row_num, 0, i, font_style)
            row_num = row_num + 1
        #======================================================



        wb.save(response)
        return response


        # qs = PayCardReport.objects.filter(report_id=report_id).order_by("shop").values()
        # data = pd.DataFrame(qs)
        # data = data.drop("report_id_id", 1)  # deleting 'report_id' column
        # data = data.drop("product", 1)  # deleting 'report_id' column
        # data = data.drop("created", 1)  # deleting 'report_id' column
        # data.set_index("id", inplace=True)
        # data.set_index("shop", inplace=True)  # sets column as titles for rows & deletes
        # data = data.T  # transposing the dataframe
        # print(data)
        # context = {
        #     "data": data.to_html(),
        # }
        # return render(request, "reports/daily_pay_card_rep.html", context)



    else:
        return render(request, "reports/daily_pay_card_rep.html")

# =====================================================================

def bonus_report_excel(request):
    users = User.objects.all()
    categories = ProductCategory.objects.all()
    doc_type = DocumentType.objects.get(name="Продажа ТМЦ")

    # start_date = request.POST["start_date"]
    # converting HTML date format (2021-07-08T01:05) to django format (2021-07-10 01:05:00)
    # start_date = datetime.strptime(start_date, "%Y-%m-%dT%H:%M")
    # end_date = request.POST["end_date"]
    # end_date = datetime.strptime(end_date, "%Y-%m-%dT%H:%M")
    rhos = RemainderHistory.objects.filter(rho_type=doc_type)

    # =======================Saving in Excel==============================================
    wb = Workbook()
    ws = wb.active
    ws.title = "Bonus"

    # ====================================End of Saving in Excel==========================

    # =====================================Array Version==========================================
    starting_row = 2
    for user in users:
        ws.cell(row=starting_row, column=1).value = user.last_name

        starting_column = 2
        for category in categories:
            ws.cell(column=starting_column, row=1).value = category.name

            if rhos.filter(user=user, category=category).exists():
                rhos = rhos.filter(user=user, category=category)
                sales = 0
                for rho in rhos:
                    sales += (
                        rho.retail_price
                        * rho.outgoing_quantity
                        * category.bonus_percent
                    )
            else:
                sales = 0

            ws.cell(column=starting_column, row=starting_row).value = sales
            starting_column += 1
        starting_row += 1

    wb.save("names.xlsx")

    return redirect("log")

def bonus_report(request):
    if request.user.is_authenticated:
        # users = User.objects.all().exclude(is_active=False)
        #users = User.objects.all()
        #users = User.objects.all().exclude(active=False)
        group_sales=Group.objects.get(name='sales')
        users = User.objects.filter(is_active=True, groups=group_sales ).order_by('username')
        categories = ProductCategory.objects.all().exclude(name='КЭО').order_by('id')
        sims=ProductCategory.objects.get(name="Сим_карты")
        shops = Shop.objects.filter(retail=True, active=True, subdealer=False)
        doc_type = DocumentType.objects.get(name="Продажа ТМЦ")
        if request.method == "POST":
            report_id = ReportTempId.objects.create()
            bulk_sim_motivation=BulkSimMotivation.objects.get(id=1)
            start_date = request.POST["start_date"]
            # converting HTML date format (2021-07-08T01:05) to django format (2021-07-10 01:05:00)
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            end_date = request.POST["end_date"]
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            end_date = end_date + timedelta(days=1)

            current_month=start_date.month
            month=Month.objects.get(id=current_month)
            current_year=start_date.year
            year=Year.objects.get(name=current_year)
        
            rhos = RemainderHistory.objects.filter(rho_type=doc_type, created__gt=start_date, created__lt=end_date)

            if Customer.objects.filter(created__gt=start_date, created__lt=end_date).exists():#look for new clients who were created in this time period
                customers=Customer.objects.filter (created__gt=start_date, created__lt=end_date)
            
            # else:
            #     messages.error(request, "Новые клиенты отсутствуют")
            #     return redirect("cashback_rep")
            if Document.objects.filter(created__gt=start_date, created__lt=end_date, title=doc_type).exists():
                documents=Document.objects.filter(created__gt=start_date, created__lt=end_date, title=doc_type)

            for user in users:
                #we create a list for each user and add "username" as the first entry
                #then we collect data for all categories in this array which later creates a row for user in excel file
                user_row = [user.username]

                #"number of work days" column
                dates=[]
                list=[]
                user_rows=rhos.filter(user=user)
                for row in user_rows:
                    #we format 'row.created' field from <class 'datetime.datetime'> to '<str>' & cut off time values in order to receive date in <str> format for further comparaison. Then we save the date in 'dates' array
                    date=row.created.strftime('%Y-%m-%d')
                    dates.append(date)
                    #we use 'numpy.unique' function to count unique values
                    list=np.unique(dates)
                number_of_wd=len(list)
                #we add number of work days as the second entry in user_row list
                user_row.append(number_of_wd)

                #cashback column
                counter=0
                if Customer.objects.filter(created__gt=start_date, created__lt=end_date, user=user).exists():
                    for customer in customers:
                        if Document.objects.filter(client=customer).exists():#search all documents within the time period fo new customer. 
                            if customer.user == user: #we add 1 to the counter even if the customer exists in more than one document
                                counter +=1
                    cash_back_bonus = counter * 15
                    user_row.append(cash_back_bonus)
                else:
                    user_row.append(0)

                #calculate bonus sum for each category for the user
                for category in categories:
                    sum = 0
                    for shop in shops:
                        rhos_new = rhos.filter(category=category, user=user, shop=shop)
                        if category.name == "Сим_карты": 
                            #checking if GI plan has been completed by the shop. We need this to chnage bonus percent
                            if KPIMonthlyPlan.objects.filter(month_reported=month, year_reported=year.name, shop=shop).exists():
                                monthly_plan=KPIMonthlyPlan.objects.get(month_reported=month, year_reported=year.name, shop=shop)
                                monthly_plan_GI=monthly_plan.GI
                                actual_GI_per_shop=rhos.filter(category=category, shop=shop, sub_total__gt=0)#исключаем нулевые сим карты
                                number_of_rhos=actual_GI_per_shop.count()
                                if monthly_plan_GI>number_of_rhos:
                                    for rho in rhos_new:
                                        #отсекаем из выручки интернет номера стоимостью > 1550 (Тариф премимум) руб
                                        if rho.sub_total <= 1550:
                                            sum += int(rho.sub_total * category.bonus_percent * shop.sale_k)
                                        else:
                                            sum += int(1550 * category.bonus_percent * shop.sale_k)
                                else:
                                    for rho in rhos_new:
                                        #отсекаем из выручки интернет номера стоимостью > 1550 (Тариф премимум) руб
                                        if rho.sub_total <= 1550:
                                            sum += int(rho.sub_total * category.bonus_percent_1 * shop.sale_k)
                                        else:
                                            sum += int(1550 * category.bonus_percent * shop.sale_k)
                            else:
                                for rho in rhos_new:
                                    sum += int(rho.sub_total * category.bonus_percent_1 * shop.sale_k)
                        else:
                            for rho in rhos_new:
                                sum += int(rho.sub_total * category.bonus_percent * shop.sale_k)
                    #add bonus sum for each category as an entry in user_row list
                    user_row.append(sum)

                #credit column
                if Credit.objects.filter(created__gt=start_date, created__lt=end_date, user=user).exists():
                    credits = Credit.objects.filter(created__gt=start_date, created__lt=end_date, user=user)
                    credit_sum = 0
                    for credit in credits:
                        credit_sum+=credit.sum
                else:
                    credit_sum = 0

                n=0
                if rhos.filter(category=sims, user=user).exists():
                    sim_rhos=rhos.filter(category=sims, user=user)
                    for rho in sim_rhos:
                        if rho.retail_price >= bulk_sim_motivation.sim_price and rho.retail_price <= 1550:
                            n+=1
                #we create a report in which categories a listed as they go in the DB
                monthly_bonus = MonthlyBonus.objects.create(
                    report_id=report_id,
                    user_name=user_row[0],
                    number_of_work_days=user_row[1],
                    cashback=user_row[2],
                    smartphones=user_row[3],
                    accessories=user_row[4],
                    sim_cards=user_row[5],
                    phones=user_row[6],
                    iphones=user_row[7],
                    insuranсе=user_row[8],
                    wink=user_row[9],
                    services=user_row[10],
                    gadgets=user_row[11],
                    modems=user_row[12],
                    RT_equipment=user_row[13],
                    protective_films=user_row[14],
                    credit=credit_sum * 0.03,
                    bulk_sims= n * bulk_sim_motivation.bonus_per_sim,
                    sub_total=0,
                )

            #==========================Convert to Excel module=========================================
            response = HttpResponse(content_type="application/ms-excel")
            response["Content-Disposition"] = (
                "attachment; filename=BonusRep_" + str(month) + str(year) + ".xls"
            )
            # str(datetime.date.today())+'.xls'

            wb = xlwt.Workbook(encoding="utf-8")
            ws = wb.add_sheet('Report')

            # sheet header in the first row
            row_num = 0
            col_num = 1
            font_style = xlwt.XFStyle()
            columns = ['Кол-во смен','Кэшбэк']
            for category in categories:
                columns.append(category.name)
            columns.append('Кредиты %')
            columns.append('Тяжелые тарифы (100)')
            for col_num in range(len(columns)):
                ws.write(row_num, col_num + 1, columns[col_num], font_style)

            # sheet body, remaining rows
            font_style = xlwt.XFStyle()
            monthly_report = MonthlyBonus.objects.filter(report_id=report_id)

            row_num = 1
            for item in monthly_report:
                col_num=0
                ws.write(row_num, col_num, item.user_name, font_style)
                col_num += 1
                ws.write(row_num, col_num, item.number_of_work_days, font_style)
                col_num += 1
                ws.write(row_num, col_num, item.cashback, font_style)
                col_num += 1
                ws.write(row_num, col_num, item.smartphones, font_style)
                col_num += 1
                ws.write(row_num, col_num, item.accessories, font_style)
                col_num += 1
                ws.write(row_num, col_num, item.sim_cards, font_style)
                col_num += 1
                ws.write(row_num, col_num, item.phones, font_style)
                col_num += 1
                ws.write(row_num, col_num, item.iphones, font_style)
                col_num += 1
                ws.write(row_num, col_num, item.insuranсе, font_style)
                col_num += 1
                ws.write(row_num, col_num, item.wink, font_style)
                col_num += 1
                ws.write(row_num, col_num, item.services, font_style)
                col_num += 1
                ws.write(row_num, col_num, item.gadgets, font_style)
                col_num += 1
                ws.write(row_num, col_num, item.modems, font_style)
                col_num += 1
                ws.write(row_num, col_num, item.RT_equipment, font_style)
                col_num += 1
                ws.write(row_num, col_num, item.protective_films, font_style)
                col_num += 1
                ws.write(row_num, col_num, item.credit, font_style)
                col_num +=1
                ws.write(row_num, col_num, item.bulk_sims, font_style)
                row_num += 1
            
            # query_set = MonthlyBonus.objects.filter().values()
            # data = pd.DataFrame(query_set)
            # data = data.drop("id", 1)
            # data = data.set_index("user_name")
            # data.to_excel("D:/Аналитика/Фин_отчет/Текущие/2021/data.xlsx")

            monthly_bonus_reports = MonthlyBonus.objects.all()
            for i in monthly_bonus_reports:
                i.delete()

            wb.save(response)
            return response

        context = {
            "categories": categories,
            #'users': users
        }
        return render(request, "reports/bonus_report.html", context)
    else:
        return redirect("login")
#======================================================================
def account_report_60_excel(request):
    if request.user.is_authenticated:
        doc_type=DocumentType.objects.get(name='Поступление ТМЦ')
        if request.method == "POST":
            start_date = request.POST["start_date"]
            end_date = request.POST["end_date"]
            # converting HTML date format (2021-07-08T01:05) to django format (2021-07-10 01:05:00)
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            tdelta = datetime.timedelta(days=1)
            end_date = end_date + tdelta
            docs=Document.objects.filter(title=doc_type, created__gte=start_date, created__lte=end_date ).order_by('created')

            #==========================Convert to Excel module=========================================
            response = HttpResponse(content_type="application/ms-excel")
            response["Content-Disposition"] = (
                "attachment; filename=_" + str(datetime.date.today()) + ".xls"
                )

            # str(datetime.date.today())+'.xls'

            wb = xlwt.Workbook(encoding="utf-8")
            ws = wb.add_sheet('60')

            # sheet header in the first row
            row_num = 0
            col_num=0
            font_style = xlwt.XFStyle()
            columns = ["Дата", 'Поставщик', 'Документ', 'Номер', 'Сумма']
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num], font_style)
                col_num =+ 1
            
           # sheet body, remaining rows
            font_style = xlwt.XFStyle()
            #docs = docs.values_list("created", "supplier", "title", "id", "sum")

            row_num = 1
            for item in docs:
                col_num = 0
                ws.write(row_num, col_num, str(item.created), font_style)
                col_num +=1
                ws.write(row_num, col_num, str(item.supplier), font_style)
                col_num +=1
                ws.write(row_num, col_num, str(item.title), font_style)
                col_num +=1
                ws.write(row_num, col_num, item.id, font_style)
                col_num +=1
                ws.write(row_num, col_num, item.sum, font_style)
                row_num +=1


            wb.save(response)
            return response
            #================================End of Excel Module========================================

            return render(request, "reports/account_report_60.html")
        

        else:
            return render(request, "reports/account_report_60.html")
    else:
        return redirect("login")
    
def account_report_62_excel(request):
    if request.user.is_authenticated:
        doc_type=DocumentType.objects.get(name='Продажа ТМЦ')
        if request.method == "POST":
            start_date = request.POST["start_date"]
            end_date = request.POST["end_date"]
            # converting HTML date format (2021-07-08T01:05) to django format (2021-07-10 01:05:00)
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            tdelta = datetime.timedelta(days=1)
            end_date = end_date + tdelta
            docs=Document.objects.filter(title=doc_type, created__gte=start_date, created__lte=end_date ).order_by('created')

            #==========================Convert to Excel module=========================================
            response = HttpResponse(content_type="application/ms-excel")
            response["Content-Disposition"] = (
                "attachment; filename=_" + str(datetime.date.today()) + ".xls"
                )

            # str(datetime.date.today())+'.xls'

            wb = xlwt.Workbook(encoding="utf-8")
            ws = wb.add_sheet('60')

            # sheet header in the first row
            row_num = 0
            col_num=0
            font_style = xlwt.XFStyle()
            columns = ["Дата", 'Покупатель', 'Документ', 'Номер', 'Сумма']
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num], font_style)
                col_num =+ 1
            
           # sheet body, remaining rows
            font_style = xlwt.XFStyle()
            #docs = docs.values_list("created", "supplier", "title", "id", "sum")

            row_num = 1
            for item in docs:
                col_num = 0
                ws.write(row_num, col_num, str(item.created), font_style)
                col_num +=1
                ws.write(row_num, col_num, str(item.client), font_style)
                col_num +=1
                ws.write(row_num, col_num, str(item.title), font_style)
                col_num +=1
                ws.write(row_num, col_num, item.id, font_style)
                col_num +=1
                ws.write(row_num, col_num, item.sum, font_style)
                row_num +=1


            wb.save(response)
            return response
            #================================End of Excel Module========================================

            return render(request, "reports/account_report_62.html")
        

        else:
            return render(request, "reports/account_report_62.html")
    else:
        return redirect("login")

def account_report_90_excel(request):
    if request.user.is_authenticated:
        doc_type=DocumentType.objects.get(name='Продажа ТМЦ')
        pay_cards=ProductCategory.objects.get(name='КЭО')
        sim_cards=ProductCategory.objects.get(name='Сим_карты')
        insurance=ProductCategory.objects.get(name='Страховки')
        subscription=ProductCategory.objects.get(name='Подписки')
        if request.method == "POST":
            start_date = request.POST["start_date"]
            end_date = request.POST["end_date"]
            # converting HTML date format (2021-07-08T01:05) to django format (2021-07-10 01:05:00)
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            tdelta = datetime.timedelta(days=1)
            end_date = end_date + tdelta
            docs=RemainderHistory.objects.filter(rho_type=doc_type, created__gte=start_date, created__lte=end_date ).exclude(category=pay_cards).exclude(category=sim_cards).exclude(category=insurance).exclude(category=subscription).order_by('created')

            #==========================Convert to Excel module=========================================
            response = HttpResponse(content_type="application/ms-excel")
            response["Content-Disposition"] = (
                "attachment; filename=_" + str(datetime.date.today()) + ".xls"
                )

            # str(datetime.date.today())+'.xls'

            wb = xlwt.Workbook(encoding="utf-8")
            ws = wb.add_sheet('60')

            # sheet header in the first row
            row_num = 0
            col_num=0
            font_style = xlwt.XFStyle()
            columns = ["Дата", 'IMEI', 'Name', 'Quantity', 'Сумма (90.1)', 'НДС (90.3)', 'Расходы на продажу (90.2)', 'Сальдо (90.9)']
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num], font_style)
                col_num =+ 1
            
           # sheet body, remaining rows
            font_style = xlwt.XFStyle()
            #docs = docs.values_list("created", "supplier", "title", "id", "sum")

            row_num = 1
            for item in docs:
                col_num = 0
                ws.write(row_num, col_num, str(item.created), font_style)
                col_num +=1
                ws.write(row_num, col_num, str(item.imei), font_style)
                col_num +=1
                ws.write(row_num, col_num, str(item.name), font_style)
                col_num +=1
                ws.write(row_num, col_num, item.outgoing_quantity, font_style)
                col_num +=1
                ws.write(row_num, col_num, item.sub_total, font_style)
                col_num +=1
                ws.write(row_num, col_num, item.sub_total/120*20, font_style)
                col_num +=1
                ws.write(row_num, col_num, item.av_price/120*100, font_style)
                col_num +=1
                ws.write(row_num, col_num, item.sub_total/120*100 - item.av_price/120*100, font_style)
                row_num +=1


            wb.save(response)
            return response
            #================================End of Excel Module========================================

            return render(request, "reports/account_report_90.html")
        

        else:
            return render(request, "reports/account_report_90.html")
    else:
        return redirect("login")
 
    if request.user.is_authenticated:
        doc_type=DocumentType.objects.get(name='Продажа ТМЦ')
        pay_cards=ProductCategory.objects.get(name='КЭО')
        sim_cards=ProductCategory.objects.get(name='Сим_карты')
        insurance=ProductCategory.objects.get(name='Страховки')
        subscription=ProductCategory.objects.get(name='Подписки')
        if request.method == "POST":
            start_date = request.POST["start_date"]
            end_date = request.POST["end_date"]
            # converting HTML date format (2021-07-08T01:05) to django format (2021-07-10 01:05:00)
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            tdelta = datetime.timedelta(days=1)
            end_date = end_date + tdelta
            docs=RemainderHistory.objects.filter(rho_type=doc_type, created__gte=start_date, created__lte=end_date ).exclude(category=pay_cards).exclude(category=sim_cards).exclude(category=insurance).exclude(category=subscription).order_by('created')

            #==========================Convert to Excel module=========================================
            response = HttpResponse(content_type="application/ms-excel")
            response["Content-Disposition"] = (
                "attachment; filename=_" + str(datetime.date.today()) + ".xls"
                )

            # str(datetime.date.today())+'.xls'

            wb = xlwt.Workbook(encoding="utf-8")
            ws = wb.add_sheet('60')

            # sheet header in the first row
            row_num = 0
            col_num=0
            font_style = xlwt.XFStyle()
            columns = ["Дата", 'Покупатель', 'Документ', 'Номер', 'Сумма']
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num], font_style)
                col_num =+ 1
            
           # sheet body, remaining rows
            font_style = xlwt.XFStyle()
            #docs = docs.values_list("created", "supplier", "title", "id", "sum")

            row_num = 1
            for item in docs:
                col_num = 0
                ws.write(row_num, col_num, str(item.created), font_style)
                col_num +=1
                ws.write(row_num, col_num, str(item.client), font_style)
                col_num +=1
                ws.write(row_num, col_num, str(item.title), font_style)
                col_num +=1
                ws.write(row_num, col_num, item.id, font_style)
                col_num +=1
                ws.write(row_num, col_num, item.sum, font_style)
                row_num +=1


            wb.save(response)
            return response
            #================================End of Excel Module========================================

            return render(request, "reports/account_report_90_2.html")
        

        else:
            return render(request, "reports/account_report_90_2.html")
    else:
        return redirect("login")

    if request.user.is_authenticated:
        doc_type=DocumentType.objects.get(name='Продажа ТМЦ')
        if request.method == "POST":
            start_date = request.POST["start_date"]
            end_date = request.POST["end_date"]
            # converting HTML date format (2021-07-08T01:05) to django format (2021-07-10 01:05:00)
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            tdelta = datetime.timedelta(days=1)
            end_date = end_date + tdelta
            docs=Document.objects.filter(title=doc_type, created__gte=start_date, created__lte=end_date ).order_by('created')

            #==========================Convert to Excel module=========================================
            response = HttpResponse(content_type="application/ms-excel")
            response["Content-Disposition"] = (
                "attachment; filename=_" + str(datetime.date.today()) + ".xls"
                )

            # str(datetime.date.today())+'.xls'

            wb = xlwt.Workbook(encoding="utf-8")
            ws = wb.add_sheet('60')

            # sheet header in the first row
            row_num = 0
            col_num=0
            font_style = xlwt.XFStyle()
            columns = ["Дата", 'Покупатель', 'Документ', 'Номер', 'Сумма']
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num], font_style)
                col_num =+ 1
            
           # sheet body, remaining rows
            font_style = xlwt.XFStyle()
            #docs = docs.values_list("created", "supplier", "title", "id", "sum")

            row_num = 1
            for item in docs:
                col_num = 0
                ws.write(row_num, col_num, str(item.created), font_style)
                col_num +=1
                ws.write(row_num, col_num, str(item.client), font_style)
                col_num +=1
                ws.write(row_num, col_num, str(item.title), font_style)
                col_num +=1
                ws.write(row_num, col_num, item.id, font_style)
                col_num +=1
                ws.write(row_num, col_num, item.sum, font_style)
                row_num +=1


            wb.save(response)
            return response
            #================================End of Excel Module========================================

            return render(request, "reports/account_report_90_3.html")
        

        else:
            return render(request, "reports/account_report_90_3.html")
    else:
        return redirect("login")